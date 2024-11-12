"""
excel_processing.py

This script processes the extracted data from 'pdf_processing' by inserting it into an XSLX file.

Author:
    Jonathan Mattis Wisser
    jmader@uni-muenster.de

Version:
    1.0
Datum:
    2024-11-12
"""

__author__ = "Jonathan Mattis Wisser"
__version__ = "1.0"
__date__ = "2024-11-12"

# ------------------------------------------------- IMPORTS ---------------------------------------------------------- #
# openpyxl zum Arbeiten mit Excel Dateien
import openpyxl

# Logging for a better understanding of the results and outputs as set up in the main module
# Logging.info() and logging.error() are used here
# https://docs.python.org/3/library/logging.html#logging.INFO
# https://docs.python.org/3/library/logging.html#logging.error
import logging


# ------------------------------------------------- UTILITY ---------------------------------------------------------- #
def find_first_empty_row(sheet):
    """
    Finds the first empty row in specified columns of an Excel worksheet to define a starting point
    for inserting the relevant information.
    This is necessary in order to be able to work as flexibly as possible with every possible Excel table

    Args:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet we want to copy the extracted information into.

    Returns:
        int: The number of the first empty row found.

    References:
        - Working with Excel worksheets using openpyxl: https://openpyxl.readthedocs.io/en/stable/api/openpyxl.worksheet.worksheet.html#openpyxl.worksheet.worksheet.Worksheet
        - Finding 'maximum column index containing data' using max_column: https://openpyxl.readthedocs.io/en/3.1/api/openpyxl.worksheet.worksheet.html#openpyxl.worksheet.worksheet.Worksheet.max_column
        - Finding 'the maximum row index containing data' using max_row: https://openpyxl.readthedocs.io/en/3.1/api/openpyxl.worksheet.worksheet.html#openpyxl.worksheet.worksheet.Worksheet.max_row
    """
    # Search for the first empty row by scanning all cells from row 1 to the maximum row + 1 for all columns for None values
    for row in range(1, sheet.max_row + 1):
        # When all cells are empty (None value) in one row for all columns, give back this row as first empty row to add new data to
        if all(sheet.cell(row=row, column=col).value is None for col in range(1, sheet.max_column + 1)):
            return row

    return sheet.max_row + 1

# ------------------------------------------------- PROCESSING ------------------------------------------------------- #
def update_excel_with_extracted_data(excel_path, extracted_data):
    """
    Updates the specified worksheet from an Excel file with the data previously extracted from the PDFs by the 'pdf_processing' module

    Args:
        excel_path (str): The full file path to the Excel file.
        extracted_data (list): List of tuples with the extracted information by the 'pdf_processing' module.

    Returns:
        workbook: The Excel file with the updated worksheet.

    References:
        - General tutorial for openpyxl: https://openpyxl.readthedocs.io/en/stable/tutorial.html
        - Working with Excel worksheets using openpyxl: https://openpyxl.readthedocs.io/en/stable/api/openpyxl.worksheet.worksheet.html
        - Loading a specific worksheet from a file: https://openpyxl.readthedocs.io/en/stable/tutorial.html#loading-from-a-file
        - Adding new data to cells for specific rows and specific columns: https://openpyxl.readthedocs.io/en/latest/tutorial.html#playing-with-data
          & https://openpyxl.readthedocs.io/en/stable/api/openpyxl.worksheet.worksheet.html#openpyxl.worksheet.worksheet.Worksheet.cell
    """

    # Open the wanted worksheet from the given Excel file
    workbook = openpyxl.load_workbook(excel_path)
    worksheet = workbook['relevantInfo']

    # Finde the row from where to start adding data by using the helper function 'find_first_empty_row()'
    start_row = find_first_empty_row(worksheet)

    # Enter the information extracted from the PDFs into the Excel file by iterating over 'extracted_data' and going one row further with each iteration
    for i, (pdf_basename, coordinates, lines_with_coordinates, drought_characterization, drought_characterization_keywords, study_type, analyzed_years, periods_with_drought, single_years_with_drought) in enumerate(extracted_data):
        # Insert the pure name of a study (or rather its PDF) into column A (Paper)
        worksheet.cell(row=start_row + i, column=1, value=pdf_basename)

        # Insert the coordinates into column B (location coordinates)
        worksheet.cell(row=start_row + i, column=2, value=coordinates)
        # Insert the study site information (either coordinate context lines or study site directly, depending what is stored in 'extracted_data') into column C (Area name)
        worksheet.cell(row=start_row + i, column=3, value=lines_with_coordinates)

        # If a time period referring to the analyzed years of a study was found, insert it into column D (time period analyzed)
        if analyzed_years:
            # Convert the list to string first, so there is no type error for the Excel file
            analyzed_years_str = ', '.join(analyzed_years)
            worksheet.cell(row=start_row + i, column=4, value=analyzed_years_str)

        # If periods or single years with drought were found, insert the combined (string) value into column E (time period with drought (if mentioned))
        if periods_with_drought or single_years_with_drought:
            # Combine both lists and convert them into one string so there is no type error for the Excel file
            combined_drought_years = periods_with_drought + single_years_with_drought
            combined_drought_years_str = ', '.join(sorted(combined_drought_years))
            worksheet.cell(row=start_row + i, column=5, value=combined_drought_years_str)

        # Insert, if a method to assess drought was found its corresponding keyword into column J (study type)
        if study_type:
            worksheet.cell(row=start_row + i, column=9, value=study_type)

        # Insert the text information, how drought was characterized into column L (how was drought characterized), if there is any
        if drought_characterization:
            worksheet.cell(row=start_row + i, column=11, value=drought_characterization)

        # Insert the found keywords of how drought was characterized into column M (drought quantification keyword for plots), if there are any
        if drought_characterization_keywords:
            # Convert the list to string first, so there is no type error for the Excel file
            drought_characterization_keywords_str = ', '.join(drought_characterization_keywords)
            worksheet.cell(row=start_row + i, column=13, value=drought_characterization_keywords_str)

    try:
        # Save the changes made in the Excel file and log that it worked
        workbook.save(excel_path)
        logging.info(f"Excel file was successfully updated!")

    # Fallback error logging if an error occurred
    except Exception as e:
        logging.error(f"Error saving the updated Excel file: {e}")